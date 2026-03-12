import time
import re
import os
import requests
from bs4 import BeautifulSoup
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
import openpyxl
from openpyxl.styles import Font

IMG_DIR = "images"
os.makedirs(IMG_DIR, exist_ok=True)

def get_driver():
    options = uc.ChromeOptions()
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1920,1080")
    driver = uc.Chrome(options=options, headless=False)
    return driver

def download_image(img_url, filename):
    try:
        headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
        r = requests.get(img_url, headers=headers, timeout=10)
        if r.status_code == 200:
            path = os.path.join(IMG_DIR, filename)
            with open(path, "wb") as f:
                f.write(r.content)
            return path
    except:
        pass
    return ""

def get_suto_links(driver):
    print("[슈퍼투데이] 목록 수집 중...")
    driver.get("https://www.suto.co.kr/cpevent")
    time.sleep(7)
    for i in range(10):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)
    soup = BeautifulSoup(driver.page_source, "html.parser")
    seen = set()
    links = []
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if not href.startswith("http"):
            href = "https://www.suto.co.kr" + href
        if "/cpevent/write" in href:
            continue
        if not re.search(r'/cpevent/\d+', href):
            continue
        if href in seen:
            continue
        seen.add(href)
        links.append(href)
    print(f"[슈퍼투데이] 유효 링크: {len(links)}건")
    return links

def parse_suto_detail(driver, url, idx):
    try:
        driver.get(url)
        time.sleep(3)

        if "cloudflare" in driver.page_source.lower() and len(driver.page_source) < 5000:
            print(f"  ✗ Cloudflare 차단: {url}")
            return None

        soup = BeautifulSoup(driver.page_source, "html.parser")

        # 제목
        title = ""
        for h1 in soup.find_all("h1"):
            t = h1.text.strip()
            if t and "이벤트응모" not in t and len(t) > 5:
                title = t
                break

        # li.no-drag 에서 정보 추출
        host = ""
        period = ""
        prize = ""
        winner_count = ""
        for li in soup.find_all("li", class_="no-drag"):
            text = li.get_text(separator=" ").strip()
            if "주최사" in text:
                host = text.replace("주최사", "").strip()
            elif "응모기간" in text:
                period = text.replace("응모기간", "").strip()
            elif "경품태그" in text:
                prize = text.replace("경품태그", "").strip()
            elif "총 당첨자수" in text:
                winner_count = text.replace("총 당첨자수", "").strip()

        # 응모하기 클릭 → 실제 원소스 URL
        external_url = url
        try:
            # requests로 리다이렉트 직접 따라가기 (브라우저 불필요)
            event_go = soup.find("a", class_="event_go")
            if event_go:
                go_href = event_go.get("href", "")
                if go_href:
                    if not go_href.startswith("http"):
                        go_href = "https://www.suto.co.kr" + go_href
                    headers = {
                        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                        "Referer": "https://www.suto.co.kr"
                    }
                    resp = requests.get(go_href, headers=headers, allow_redirects=True, timeout=10)
                    external_url = resp.url  # 최종 리다이렉트 URL
                    print(f"     → 원소스: {external_url[:60]}")
        except:
            external_url = go_href if go_href else url

        ## 썸네일
        img_path = ""
        try:
            # poster-container 안에 있음
            poster = soup.find("div", class_="poster-container")
            if poster:
                img_el = poster.find("img")
            else:
                img_el = None

            if img_el:
                img_url = img_el.get("src", "")
                if img_url and not img_url.startswith("http"):
                    img_url = "https://www.suto.co.kr" + img_url
                if img_url and img_url.startswith("http"):
                    ext = img_url.split(".")[-1].split("?")[0][:4]
                    if ext not in ["jpg", "jpeg", "png", "gif", "webp"]:
                        ext = "jpg"
                    filename = f"suto_{idx:04d}.{ext}"
                    img_path = download_image(img_url, filename)
        except:
            pass

        print(f"  ✓ {title[:25]} | 주최: {host} | 경품: {prize[:20]} | 기간: {period}")

        return {
            "출처": "슈퍼투데이",
            "제목": title,
            "주최사": host,
            "응모기간": period,
            "경품": prize,
            "당첨자수": winner_count,
            "응모링크": external_url,
            "슈퍼투데이링크": url,
            "썸네일경로": img_path
        }

    except Exception as e:
        print(f"  ✗ 오류: {url} - {e}")
        return None

def save_excel(results, filename="딥크롤링결과.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "슈퍼투데이"
    headers = ["출처", "제목", "주최사", "응모기간", "경품", "당첨자수", "응모링크", "슈퍼투데이링크", "썸네일경로"]
    ws.append(headers)
    for row in results:
        ws.append([
            row.get("출처", ""),
            row.get("제목", ""),
            row.get("주최사", ""),
            row.get("응모기간", ""),
            row.get("경품", ""),
            row.get("당첨자수", ""),
            row.get("응모링크", ""),
            row.get("슈퍼투데이링크", ""),
            row.get("썸네일경로", ""),
        ])
        r = ws.max_row
        for col, key in [(7, "응모링크"), (8, "슈퍼투데이링크")]:
            cell = ws.cell(row=r, column=col)
            link = row.get(key, "")
            if link and link.startswith("http"):
                cell.hyperlink = link
                cell.font = Font(color="0000FF", underline="single")
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
    wb.save(filename)
    print(f"\n✅ 저장 완료: {filename} ({len(results)}건)")

if __name__ == "__main__":
    driver = get_driver()
    results = []
    try:
        links = get_suto_links(driver)
        links = links[:10]  # ← 이 줄 추가! 10개만
        print(f"\n[딥크롤링] 상세페이지 {len(links)}건 시작...")
        for idx, url in enumerate(links):
            item = parse_suto_detail(driver, url, idx)
            if item:
                results.append(item)
            time.sleep(2)
    finally:
        driver.quit()
    save_excel(results)